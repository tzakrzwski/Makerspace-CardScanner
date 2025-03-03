import sys
import customtkinter as ctk
from openpyxl import load_workbook
import tkinter as tk
from tkinter import simpledialog, Canvas, messagebox
from PIL import Image, ImageTk 
from screeninfo import get_monitors
import pygetwindow as gw
from datetime import datetime
from __main__ import *
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException

#Path to excel sheet
file_path = "hardware_users.xlsx"
sheet_name = "Scans"
sheet2_name ="Users"
Location= "Watt"

def load_excel(): #this is a bit redundant but it works, can eventually use this in "add user to shee"
    # Load the workbook and sheet
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name] 
    sheet2 = workbook[sheet2_name] 
    return workbook, sheet, sheet2

#def load_hardware_ids(sheet): #Not even using this rn
    hardware_dict = {}
    # Load the hardware IDs and usernames into a dictionary
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        hardware_dict[str(row[0])] = row[1]  # Hardware ID as the key, username as the value

def find_hardware_id(sheet2, hardware_id):
    # Loop through the rows starting from row 2 to skip headers (if any)
    for row in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, values_only=True):
        if str(row[1]) == str(hardware_id):  # Compare hardware_id in column B (index 1)
            return row[0]  # Return the username from column A (index 0)
    
    return None  # Return None if not found
def find_userdata(hardware_id,sheet2):
    # Loop through the rows starting from row 2 to skip headers
    for row in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, values_only=True):
        if str(row[1]) == str(hardware_id):  # Look for hardware_id in col. B
            first_name = row[3]  # Column D 
            last_name = row[4]   # Column E 
            major = row[5]       # Column F
            return first_name, last_name, major
    
    return None, None, None  #   Set to None if they are not found

def add_user_to_sheet(sheet_name,sheet2_name,hardware_id, username,first_name,last_name,major,workbook,userstatus):
    wb = workbook
    scans_sheet = wb[sheet_name]
    users_sheet = wb[sheet2_name]
    
    if userstatus == 1:
            # Search for matching hardware ID in the "Users" sheet or for an empty hardware ID cell
            for row in users_sheet.iter_rows(min_row=2, values_only=False):  # Skip header row
                cell_hardware_id = row[1].value  # Column B in "Users" for hardware_id

            # Cast both the hardware_id from input and the one from the sheet to str for comparison
                if str(cell_hardware_id) == str(hardware_id): #str is irrelavent but may help edge cases where numbers get input as strings?
                    match_found = True
                    print(f"User with hardware ID {hardware_id} already exists in 'Users' sheet.")
                    break  # Stop searching after finding the match

                # If the hardware ID cell is empty (i.e., new entry row), fill in this row
                if cell_hardware_id is None or cell_hardware_id == "":  # Check for an empty hardware ID
                    row[0].value = username  # Column A for username
                    row[1].value = int(hardware_id)  # Column B for hardware ID
                    row[3].value = first_name  # Column D for first name
                    row[4].value = last_name   # Column E for last name
                    row[5].value = major       # Column F for major
                    match_found = True
                    print(f"New user {first_name} {last_name} added to 'Users' sheet.")
                    break  # Stop searching after appending the new data

    # Add the scan to the 'Scans' sheet (this happens regardless of userstatus)
    now = datetime.now()
    #timestamp = now.timestamp() #use this for seconds only
    timestamp = now.strftime('%m/%d/%Y %H:%M:%S') # Format the time to display as "YYYY-MM-DD HH:MM"
    scans_sheet.append([int(hardware_id), username, timestamp])
    
    # Save the workbook after making changes
    wb.save(file_path)
    print(f"Scan Added, workbook saved.")


def scrape_user(username):
    # Set up Selenium with headless Chrome
        chrome_options = Options()
        chrome_options.add_argument("--headless")  # Run Chrome in headless mode
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")

        # Set up the driver
        driver = webdriver.Chrome(options=chrome_options)

        # Create da URL using username, change this id the directory changes url
        url = f"https://my.clemson.edu/#/directory/person/{username}"

        # Load da page
        driver.get(url)

        # Wait for the full name element to appear
        try:
            WebDriverWait(driver, 1).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '.personView .primaryInfo h2')) 
            )
            #the time should be upped if the scrape does not complete in time
        except TimeoutException:
            print(f"Timeout while waiting for the element on the page for {username}.")
            driver.quit()
            return username, None, None  # Default values if element not found
        except Exception as e:
            print(f"Error during page load: {e}")
            driver.quit()
            return username, None, None  # Default values if there's an error
        # Get da page
        page_source = driver.page_source

        # Parse the page with BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        #Find da Name
        full_name_element = soup.select_one('.personView .primaryInfo h2') 
        if full_name_element:
            full_name = full_name_element.get_text().strip()

            # Split the full name by spaces
            name_parts = full_name.split()

            # Store the first name and last name (ignore middle name if present)
            first_name = name_parts[0]  # The first part of the name
            last_name = name_parts[-1]  # The last part of the name, if they have like a number or sr. or something after their last name itll fuck it up but it doesn't matter because we don't really need the last name that badly anyways
        else:
            first_name, last_name = None, None
            
        #Find major
        major_element = soup.select_one('.personView .primaryInfo .data p')
        if major_element:
            major = major_element.get_text().strip()
        else:
            major = None
        # Print the scraped information
        print(f"First Name: {first_name}")
        print(f"Last Name: {last_name}")
        print(f"Major: {major}")

        # Close the scraping driver/borderless window
        driver.quit()
        return first_name, last_name, major

def make_fullscreen_on_top(root):
    root.attributes('-fullscreen', True)
    root.attributes('-topmost', True)
#this can probably be put within another function but it works for now

def show_welcome_popup(root, username, first_name, userstatus):
    # Set the background image
    image = Image.open("background.png")
    bg_image = ImageTk.PhotoImage(image)
    root.bg_image = bg_image  # Keep a reference to avoid garbage collection
    #background_label = tk.Label(root, image=bg_image)   # Create a label for the background
    #background_label.place(relwidth=1, relheight=1)  # Stretch to fit window (Idek if this works properly because it isn't doing it that well)
    
    # welcome back message
    if first_name == None:
        first_name = username
    if userstatus == 0:
        message = f"Welcome back, {first_name}!"
    else:
        message = f"Welcome to the {Location} Makerspace!"
    
    message_label = tk.Label(root, text=message, font=("Vendetta", 50, "bold"), fg="white", bg="black")
    message_label.place(relx=0.5, rely=0.5, anchor="center")  # Center the message


    ### Not using this currently because tkinter doesn't seem to allow for transparent stacking of images
    # Load and place the logo below the text
    #logo_image = Image.open("LogoBW.png")  # Replace with your logo image file
    #logo_image = logo_image.resize((508, 128))  # Resize if needed
    #logo_photo = ImageTk.PhotoImage(logo_image)
    #logo_label = tk.Label(root, image=logo_photo)
    #logo_label.place(relx=0.5, rely=0.7, anchor="center")

    root.after(3000, root.quit)  # Close the window after 3 seconds

def close_on_escape(event): #this is maybe redundant because I do this inside some of the definitions
    print("Escape key pressed. Exiting program...") #esc to close
    sys.exit()  # Exit the program

def prompt_for_username():
    # Simple dialog to ask for a username
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    def is_valid_username(entered_username):
        """Validate the username based on the provided rules."""
        if not entered_username.strip():
            return False, "Username cannot be blank."
        
        if entered_username.isdigit():
            return False, "Username cannot be all numbers."
        
        if "@" in entered_username and "." in entered_username:
            return False, "Username cannot be an email address."
        
        return True, "Valid username."
    
    def submit_username(event=None):
        entered_username = entry.get()
        valid, message = is_valid_username(entered_username)
        
        if valid:
            nonlocal username  # Declare nonlocal to update username within the nested function
            username = entered_username  # Store the valid username
            root.quit()  # Stop the local mainloop
            root.destroy()  # Close the username entry window
        else:
            messagebox.showerror("Error", message)
    
    # Set up the username variable to hold the result
    username = None
    
    # Initialize the main window
    ctk.set_appearance_mode("dark")  # Modes: "dark" or "light"
    ctk.set_default_color_theme("blue")  # We will override the default colors manually
    
    root = ctk.CTk()
    root.title("Username Entry")
    root.geometry("1500x500")  # Setting a larger window size
    
    # Override the color theme to use the clemson orange
    orange_color = "#F56600"
    
    # Title label
    title_label = ctk.CTkLabel(master=root, text="Welcome to the Makerspace! Enter Your Clemson Username:", font=("Arial", 24), text_color=orange_color)
    title_label.pack(pady=30)
    
    # Create a label and entry for the username
    label = ctk.CTkLabel(master=root, text="The part before the @clemson.edu", font=("Arial", 18))
    label.pack(pady=10)
    
    entry = ctk.CTkEntry(master=root, width=300, height=50, placeholder_text="Enter username", font=("Arial", 16))
    entry.pack(pady=10)
    entry.focus() #trying to focus on the text box
   
    # Create a submit button with the orange color
    submit_button = ctk.CTkButton(master=root, text="Submit", command=submit_username, width=200, height=50, fg_color=orange_color, hover_color="#FF7800", font=("Arial", 18))
    submit_button.pack(pady=40)
    
    # Bind the Enter key to submit the form
    root.bind('<Return>', submit_username)  # Bind Enter (Return) key to the submit function
    
    # Start the application loop
    root.mainloop()
    # Return the username after the window closes
    return username

def main(hardware_id=None):
    userstatus=None
    if hardware_id == None:
        hardware_id = sys.argv[1] #This gets the hardware ID from the gloabl system variables as defined from the other script to pass along the variables.
    workbook,sheet,sheet2 = load_excel()
    username = find_hardware_id(sheet2, hardware_id)
    first_name=None
    major=None
    root = tk.Tk()
    root.withdraw()  # Hide the root window initially
    root.bind("<Escape>", close_on_escape) # Bind the Escape key to close the program

    if username != None:
        print(f"User found: {username}")
        userstatus=0
        first_name,last_name,major = find_userdata(hardware_id, sheet2)
        add_user_to_sheet(sheet_name,sheet2_name,hardware_id, username,first_name,last_name,major,workbook,userstatus)
        show_welcome_popup(root,username,first_name,userstatus)
        root.deiconify()  # Show the window
        make_fullscreen_on_top(root)
        root.mainloop()
    else:
        print("New user detected. Prompting for username.")
        username = prompt_for_username()
        userstatus=1
        show_welcome_popup(root,username,first_name,userstatus)
        first_name, last_name, major = scrape_user(username)
        add_user_to_sheet(sheet_name,sheet2_name,hardware_id, username,first_name,last_name,major,workbook,userstatus)
        workbook.save(file_path)
        username=None
            
    
if __name__ == "__main__":
    main()